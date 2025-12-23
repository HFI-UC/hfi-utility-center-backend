from collections import defaultdict
from typing import Literal, Sequence
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from playwright.async_api import async_playwright
from core.env import *
from core.orm import *
from core.email import *
import secrets
import uuid
import httpx
import hashlib
import hmac
import time
import json


def get_exported_xlsx(
    reservations: Sequence[Reservation],
    format: Literal["by-room", "single-sheet"] = "by-room",
) -> Workbook:
    os.makedirs("cache", exist_ok=True)
    workbook = Workbook()
    default = workbook.active
    if default is not None:
        workbook.remove(default)

    headers = [
        "ID",
        "Start Time",
        "End Time",
        "Student Name",
        "Student ID",
        "E-mail",
        "Reason",
        "Room Name",
        "Class Name",
        "Status",
        "Creation Time",
        "Campus Name",
    ]

    if format == "single-sheet":
        ws: Worksheet = workbook.create_sheet(title="All Reservations")
        ws.append(headers)

        for reservation in reservations:
            room = reservation.room
            class_ = reservation.class_
            campus = room.campus if room and room.campus else None
            ws.append(
                [
                    reservation.id,
                    reservation.startTime,
                    reservation.endTime,
                    reservation.studentName,
                    reservation.studentId,
                    reservation.email,
                    reservation.reason,
                    room.name if room else None,
                    class_.name if class_ else None,
                    reservation.status.capitalize() if reservation.status else None,
                    reservation.createdAt,
                    campus.name if campus else None,
                ]
            )
    else:
        reservations_by_room: dict[int | None, list[Reservation]] = defaultdict(list)
        for reservation in reservations:
            reservations_by_room[reservation.roomId].append(reservation)

        used = set()
        for room_id, room_reservations in reservations_by_room.items():
            room_obj = next((res.room for res in room_reservations if res.room), None)
            room_name = (room_obj.name if room_obj else None) or f"Room-{room_id}"
            base = (room_name or "")[:31]
            sheet_name = base
            i = 1
            while sheet_name in used:
                suffix = f"-{i}"
                if len(base) + len(suffix) > 31:
                    sheet_name = base[: 31 - len(suffix)] + suffix
                else:
                    sheet_name = base + suffix
                i += 1
            used.add(sheet_name)

            ws: Worksheet = workbook.create_sheet(title=sheet_name)
            ws.append(headers)

            for reservation in room_reservations:
                room = reservation.room
                class_ = reservation.class_
                campus = room.campus if room and room.campus else None
                ws.append(
                    [
                        reservation.id,
                        reservation.startTime,
                        reservation.endTime,
                        reservation.studentName,
                        reservation.studentId,
                        reservation.email,
                        reservation.reason,
                        room.name if room else None,
                        class_.name if class_ else None,
                        reservation.status.capitalize() if reservation.status else None,
                        (reservation.createdAt if reservation.createdAt else None),
                        campus.name if campus else None,
                    ]
                )
    return workbook


def verify_turnstile_token(token: str) -> bool:
    try:
        with httpx.Client() as client:
            response = client.post(
                "https://challenges.cloudflare.com/turnstile/v0/siteverify",
                json={"secret": cloudflare_secret, "response": token},
            )
            data = response.json()
            if data["success"]:
                return True
            else:
                return False
    except Exception:
        return False


async def get_exported_pdf(url: str, output: str, device_scale: int = 2) -> None:
    os.makedirs("cache", exist_ok=True)
    async with async_playwright() as p:
        browser = await p.chromium.launch()
        context = await browser.new_context(
            viewport={"width": 800, "height": 900},
            device_scale_factor=device_scale,
        )
        page = await context.new_page()
        await page.goto(url, wait_until="networkidle")
        await page.emulate_media(media="print")
        await page.wait_for_timeout(2000)
        await page.pdf(
            path=output,
            format="A4",
            print_background=True,
            prefer_css_page_size=True,
            margin={"bottom": "6mm", "top": "6mm"},
        )
        await browser.close()


async def get_screenshot(url: str, output: str, device_scale: int = 2) -> None:
    os.makedirs("cache", exist_ok=True)
    async with async_playwright() as p:
        browser = await p.chromium.launch()
        context = await browser.new_context(
            viewport={"width": 800, "height": 900},
            device_scale_factor=device_scale,
        )
        page = await context.new_page()
        await page.goto(url, wait_until="networkidle")
        await page.emulate_media(media="print")
        await page.wait_for_timeout(2000)
        await page.screenshot(
            path=output,
            full_page=True,
        )
        await browser.close()


async def ai_approval(session: AsyncSession, id: int) -> None:
    reservation = await get_reservation_by_id(session, id)
    if not reservation:
        return
    async with httpx.AsyncClient(timeout=10000) as client:
        response = await client.get(
            ai_approval_url,
            params={"s": ai_approval_secret, "reason": reservation.reason},
        )
        response_json = response.json()
        data = AIApprovalResponse.model_validate(response_json)
        if data.status != "pending":
            await change_reservation_status_by_id(
                session,
                reservation.id,
                "approved" if data.status == "approved" else "rejected",
                ai_approval_admin_id,
                data.message,
            )
        if data.status == "approved":
            send_reservation_approval_email(
                email_title="Reservation Approval",
                title="Your reservation has been approved!",
                email=reservation.email,
                details=f"Hi {reservation.studentName}! Your reservation #{reservation.id} for {reservation.room.name if reservation.room else None} has been approved. Below is the detailed information.",
                user=reservation.studentName,
                room=reservation.room.name if reservation.room else "",
                class_name=reservation.class_.name or "",
                student_id=reservation.studentId,
                reason=reservation.reason,
                time=f"{reservation.startTime.strftime('%Y-%m-%d %H:%M')} - {reservation.endTime.strftime('%H:%M')}",
            )
        elif data.status == "rejected":
            send_normal_update_email(
                email_title="Reservation Rejected",
                title="Your reservation has been rejected.",
                email=reservation.email,
                details=f"Hi {reservation.studentName}! Your reservation #{reservation.id} for {reservation.room.name if reservation.room else None} has been rejected. Reason: {data.message}",
            )
        else:
            for approver in reservation.room.approvers:
                user = approver.user
                if not user or not approver.notificationsEnabled:
                    continue
                token = secrets.token_hex(32)
                send_normal_update_with_external_link_email(
                    email_title="New Reservation Request",
                    title=f"Hi {user.name}! A new reservation request has been created.",
                    email=user.email,
                    details=f"Reservation ID #{reservation.id}, click the button below for reservation details.",
                    button_text="View Reservation",
                    link=f"{base_url}/admin/reservation/?token={token}",
                )
                await create_login_token(session, user.email, token)


def get_temp_cos_security_token(ext: str) -> COSCredentialsResponse | None:
    def generate_cos_key(ext: str) -> str:
        file_name = f"{uuid.uuid4()}{ext if ext else ''}"
        return f"file/{datetime.now().strftime('%Y%m%d')}/{file_name}"

    key = generate_cos_key(ext)
    resource = (
        f"qcs::cos:{cos_region}:uid/{str(cos_bucket).split('-')[1]}:{cos_bucket}/{key}"
    )
    policy = {
        "version": "2.0",
        "statement": [
            {
                "action": [
                    "name/cos:PutObject",
                    "name/cos:InitiateMultipartUpload",
                    "name/cos:ListMultipartUploads",
                    "name/cos:ListParts",
                    "name/cos:UploadPart",
                    "name/cos:CompleteMultipartUpload",
                ],
                "effect": "allow",
                "resource": [resource],
                "condition": {
                    "string_like": {"cos:content-type": "image/*"},
                    "numeric_less_than_equal": {"cos:content-length": 5 * 1024 * 1024},
                },
            }
        ],
    }

    action = "GetFederationToken"
    version = "2018-08-13"
    algorithm = "TC3-HMAC-SHA256"
    timestamp = int(time.time())
    date = datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d")

    params = {
        "DurationSeconds": 180,
        "Name": "hfi-utility-center-temp-token",
        "Policy": json.dumps(policy, separators=(",", ":")),
    }
    payload = json.dumps(params, separators=(",", ":"))

    ct = "application/json; charset=utf-8"
    canonical_headers = f"content-type:{ct}\nhost:sts.tencentcloudapi.com\nx-tc-action:{action.lower()}\n"
    signed_headers = "content-type;host;x-tc-action"
    hashed_request_payload = hashlib.sha256(payload.encode("utf-8")).hexdigest()
    canonical_request = (
        f"POST\n/\n\n{canonical_headers}\n{signed_headers}\n{hashed_request_payload}"
    )

    credential_scope = f"{date}/sts/tc3_request"
    hashed_canonical_request = hashlib.sha256(
        canonical_request.encode("utf-8")
    ).hexdigest()
    string_to_sign = (
        f"{algorithm}\n{timestamp}\n{credential_scope}\n{hashed_canonical_request}"
    )

    def sign(key, msg):
        return hmac.new(key, msg.encode("utf-8"), hashlib.sha256).digest()

    secret_date = sign(("TC3" + cos_secret_key).encode("utf-8"), date)
    secret_service = sign(secret_date, "sts")
    secret_signing = sign(secret_service, "tc3_request")
    signature = hmac.new(
        secret_signing, string_to_sign.encode("utf-8"), hashlib.sha256
    ).hexdigest()

    authorization = (
        f"{algorithm} "
        f"Credential={cos_secret_id}/{credential_scope}, "
        f"SignedHeaders={signed_headers}, "
        f"Signature={signature}"
    )

    headers = {
        "Content-Type": ct,
        "Host": "sts.tencentcloudapi.com",
        "X-TC-Action": action,
        "X-TC-Version": version,
        "X-TC-Timestamp": str(timestamp),
        "X-TC-Region": cos_region,
        "Authorization": authorization,
    }

    try:
        with httpx.Client() as client:
            response = client.post(
                "https://sts.tencentcloudapi.com/",
                headers=headers,
                content=payload,
            )
            data = response.json()
            print(data)
            if "Response" not in data or "Credentials" not in data["Response"]:
                return None
            return COSCredentialsResponse(
                **{
                    **data["Response"]["Credentials"],
                    "StartTime": timestamp,
                    "ExpiredTime": data["Response"]["ExpiredTime"],
                    "Key": key,
                    "Bucket": cos_bucket,
                    "Region": cos_region,
                }
            )
    except Exception as e:
        print(f"Exception getting token: {e}")
        return None
